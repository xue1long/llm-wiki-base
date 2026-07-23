# ruflo-kb/src/utils/extract/office.py
"""Office document text extraction (DOCX, XLSX).

Distinguishes two failure modes that callers care about:

* ``UnsupportedFormat`` — the extension is one we cannot extract at all
  (e.g. legacy ``.doc`` binary OLE Compound File, which ``python-docx``
  cannot read).
* ``EncryptedDocumentError`` — the file is encrypted, password-protected,
  or is not a valid document of its declared type (corrupt DOCX / XLSX).
"""
import zipfile
from pathlib import Path

from .exceptions import EncryptedDocumentError, UnsupportedFormat


def extract_office_text(file_path: str) -> str:
    """从 Office 文档提取文本"""
    ext = Path(file_path).suffix.lower()

    if ext == ".doc":
        # Legacy binary OLE Compound File format — python-docx cannot read it.
        # Surface a clear, typed error so callers can prompt the user to
        # convert the file to .docx first.
        raise UnsupportedFormat(
            "Legacy .doc format is not supported; convert the file to .docx first"
        )

    if ext == ".docx":
        return extract_docx_text(file_path)
    elif ext in (".xlsx", ".xls"):
        return extract_xlsx_text(file_path)
    else:
        raise ValueError(f"Unsupported office format: {ext}")


def extract_docx_text(file_path: str) -> str:
    """从 DOCX 提取文本"""
    try:
        from docx import Document
    except ImportError as exc:  # pragma: no cover - python-docx is optional
        raise EncryptedDocumentError(
            "python-docx is not installed; cannot extract DOCX text"
        ) from exc

    try:
        doc = Document(file_path)
    except Exception as exc:
        # python-docx surfaces corrupt / encrypted / password-protected DOCX
        # files as PackageNotFoundError or zipfile.BadZipFile. Wrap them in
        # a typed error so callers can distinguish "encryption" from generic
        # IO failures.
        _raise_if_encrypted(exc, suffix=".docx")
        raise

    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def extract_xlsx_text(file_path: str) -> str:
    """从 XLSX 提取文本"""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - openpyxl is optional
        raise EncryptedDocumentError(
            "openpyxl is not installed; cannot extract XLSX text"
        ) from exc

    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as exc:
        _raise_if_encrypted(exc, suffix=".xlsx")
        raise

    parts = []

    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                parts.append("\t".join(cells))

    return "\n".join(parts)


def _raise_if_encrypted(exc: Exception, suffix: str) -> None:
    """Inspect the exception and raise EncryptedDocumentError if appropriate.

    Covers:
      * ``zipfile.BadZipFile`` — file is not a valid Office document
      * ``docx.opc.exceptions.PackageNotFoundError`` — DOCX package missing
      * ``cryptocode.Unknown`` / decryption errors from python-docx
    """
    name = type(exc).__name__
    msg = str(exc).lower()

    # Cryptocode-style decryption errors from python-docx
    if name == "PackageNotFoundError" or "package not found" in msg:
        raise EncryptedDocumentError(
            f"{suffix} file is corrupt or password-protected (PackageNotFoundError)"
        ) from exc

    # zipfile.BadZipFile — file is not a valid zip-based Office document
    if isinstance(exc, zipfile.BadZipFile):
        raise EncryptedDocumentError(
            f"{suffix} file is not a valid Office document (corrupt or encrypted)"
        ) from exc

    # Heuristic: python-docx raises a generic Exception with "decrypt" /
    # "password" / "encrypted" in the message for encrypted files.
    if any(token in msg for token in ("decrypt", "password", "encrypted")):
        raise EncryptedDocumentError(
            f"{suffix} file is encrypted or password-protected"
        ) from exc
