# ruflo-kb/src/collector/converter/office_converter.py
"""DOCX / XLSX → Markdown 转换器。

复用 utils/extract/office.py 的文本提取，加上结构化 Markdown 格式化。
"""
from __future__ import annotations

import tempfile
from pathlib import Path

from .base import ConvertResult, ConverterBase
from ...utils.extract.office import extract_docx_text, extract_xlsx_text


class OfficeConverter(ConverterBase):
    """DOCX / XLSX → 结构化 Markdown。"""

    _EXTS = {".docx", ".xlsx", ".xls"}

    def can_handle(self, source: str | Path) -> bool:
        return Path(str(source)).suffix.lower() in self._EXTS

    async def convert(
        self,
        source: str | Path,
        *,
        content: bytes | None = None,
    ) -> ConvertResult:
        source_str = str(source)
        ext = Path(source_str).suffix.lower()

        # bytes 输入 → 写临时文件
        tmp_path: str | None = None
        if content is not None:
            tmp = tempfile.NamedTemporaryFile(suffix=ext, delete=False)
            tmp.write(content)
            tmp.close()
            tmp_path = tmp.name
            source_str = tmp_path

        try:
            if ext == ".docx":
                text = extract_docx_text(source_str)
            else:
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    text = extract_xlsx_text(source_str)
                else:
                    wb = load_workbook(source_str, data_only=True)
                    text = self._xlsx_to_markdown_table(wb)
        finally:
            if tmp_path:
                Path(tmp_path).unlink(missing_ok=True)

        title = Path(str(source)).stem
        md = f"# {title}\n\n{text}"

        return ConvertResult(
            content=md,
            title=title,
            metadata={"format": ext},
            source_type=ext.lstrip("."),
            original_path=str(source),
        )

    @staticmethod
    def _xlsx_to_markdown_table(wb) -> str:
        """将 openpyxl Workbook 转为 Markdown 表格。"""
        parts: list[str] = []
        for ws in wb.worksheets:
            parts.append(f"\n## {ws.title}\n")
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                parts.append("(空工作表)")
                continue
            header = [str(c) if c is not None else "" for c in rows[0]]
            parts.append("| " + " | ".join(header) + " |")
            parts.append("| " + " | ".join(["---"] * len(header)) + " |")
            for row in rows[1:]:
                cells = [str(c) if c is not None else "" for c in row]
                while len(cells) < len(header):
                    cells.append("")
                parts.append("| " + " | ".join(cells[: len(header)]) + " |")
        return "\n".join(parts)
